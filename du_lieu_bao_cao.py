import os
import sys
import subprocess
import logging
import pandas as pd
from playwright.sync_api import sync_playwright, Page
import pytest
from openpyxl import load_workbook
from datetime import datetime

# Phần code cài đặt trình duyệt và fixtures Pytest giữ nguyên
try:
    from playwright._impl._installer import install

    install("chromium")
except Exception:
    try:
        subprocess.run(
            [sys.executable, "-m", "playwright", "install", "chromium"],
            check=True
        )
    except Exception as e:
        print("Không thể tải Chromium:", e)
        sys.exit(1)

if getattr(sys, 'frozen', False):
    BASE_DIR = sys._MEIPASS
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))


@pytest.fixture(scope="session")
def browser():
    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=False,
            args=["--disable-blink-features=AutomationControlled", "--disable-animations", "--start-maximized"]
        )
        yield browser
        browser.close()


@pytest.fixture
def page(browser):
    context = browser.new_context(no_viewport=True)
    page = context.new_page()
    yield page
    context.close()


# --- Test Case Chính đã sửa ---
def test_lay_thong_tin_du_an(page: Page):
    excel_path = os.path.join(BASE_DIR, "data.xlsx")

    # Sửa: Đọc file Excel, bỏ qua header để lấy danh sách project từ hàng 2
    project_df = pd.read_excel(excel_path, sheet_name="BaoCao", header=None)
    # Sửa: Lấy danh sách từ hàng thứ 2 (chỉ số 1) trở đi của cột đầu tiên (chỉ số 0)
    project_list = project_df.iloc[1:, 0].tolist()

    wb = load_workbook(excel_path)
    ws = wb["BaoCao"]

    # 1. Đăng nhập
    page.goto("https://qlvh.khaservice.com.vn/login")
    page.locator("input[name='email']").fill("admin@khaservice.com.vn")
    page.locator("input[name='password']").fill("Kha@@123")
    page.locator("button[type='submit']").click()
    page.wait_for_timeout(2000)

    # 2. Vòng lặp cập nhật danh mục
    for idx, project_val in enumerate(project_list, start=2):  # Sửa: Bắt đầu idx từ 2
        print(f"[{idx}] Project={project_val}")
        logging.error(f"[{idx}] - Project={project_val}")

        page.locator("#combo-box-demo").click()
        page.locator("#combo-box-demo").fill(str(project_val))
        page.locator("#combo-box-demo-option-0").click()

        page.locator("a[href='/statistics/overview']").click()
        page.wait_for_timeout(500)

        # Sửa lỗi cú pháp XPath
        tong_can_ho = page.locator('//*[@id="root"]/div[2]/main/div/div/div/div[2]/div/div[1]/p[1]').inner_text()
        tong_cu_dan = page.locator('//*[@id="root"]/div[2]/main/div/div/div/div[3]/div/div[1]/p[1]').inner_text()
        tong_cu_dan_su_dung_app = page.locator(
            '//*[@id="root"]/div[2]/main/div/div/div/div[5]/div/div[1]/p[1]').inner_text()
        tong_can_ho_su_dung_app = page.locator(
            '//*[@id="root"]/div[2]/main/div/div/div/div[6]/div/div[1]/p[1]').inner_text()

        # Ghi các giá trị vào các cột B, C, D, E của hàng tương ứng với idx
        ws[f"B{idx}"] = tong_can_ho
        ws[f"C{idx}"] = tong_cu_dan
        ws[f"D{idx}"] = tong_cu_dan_su_dung_app
        ws[f"E{idx}"] = tong_can_ho_su_dung_app

    # Lưu file
    wb.save(excel_path)
    print("Đã ghi xong dữ liệu vào file Excel.")
    page.close()

def test_lay_so_luong_bai_viet_loai_tin_tuc(page: Page):
    excel_path = os.path.join(BASE_DIR, "data.xlsx")
    base_url = "https://qlvh.khaservice.com.vn"
    # Sửa: Đọc file Excel, bỏ qua header để lấy danh sách project từ hàng 2
    project_df = pd.read_excel(excel_path, sheet_name="BaoCao", header=None)
    # Sửa: Lấy danh sách từ hàng thứ 2 (chỉ số 1) trở đi của cột đầu tiên (chỉ số 0)
    project_list = project_df.iloc[1:, 0].tolist()

    wb = load_workbook(excel_path)
    ws = wb["BaoCao"]

    # 1. Đăng nhập
    page.goto("https://qlvh.khaservice.com.vn/login")
    page.locator("input[name='email']").fill("admin@khaservice.com.vn")
    page.locator("input[name='password']").fill("Kha@@123")
    page.locator("button[type='submit']").click()
    page.wait_for_timeout(2000)

    page.goto(f"{base_url}/posts/news")
    page.locator("//*[@id='root']/div[2]/main/div/div/div[3]/div/div[2]/button").click()
    page.locator("//*[@id='menu-apartment-list-style1']/div[3]/ul/li[6]").click()
    page.wait_for_timeout(2000)

    # 2. Vòng lặp cập nhật danh mục
    for idx, project_val in enumerate(project_list, start=2):  # Sửa: Bắt đầu idx từ 2
        print(f"[{idx}] Project={project_val}")
        logging.error(f"[{idx}] - Project={project_val}")

        page.locator("#combo-box-demo").click()
        page.locator("#combo-box-demo").fill(str(project_val))
        page.locator("#combo-box-demo-option-0").click()

        page.wait_for_timeout(1000)
        rows = page.locator('//*[@id="root"]/div[2]/main/div/div/div[2]/table/tbody/tr')
        tin_tuc_count = rows.count()

        logging.error(f"[{idx}] - Project:{project_val} - Tin tuc:{tin_tuc_count}")

        ws[f"F{idx}"] = tin_tuc_count
    wb.save(excel_path)
    print("Đã ghi xong dữ liệu vào file Excel.")
    page.close()

def test_lay_so_luong_bai_viet_loai_thong_bao(page: Page):
    excel_path = os.path.join(BASE_DIR, "data.xlsx")
    base_url = "https://qlvh.khaservice.com.vn"
    # Sửa: Đọc file Excel, bỏ qua header để lấy danh sách project từ hàng 2
    project_df = pd.read_excel(excel_path, sheet_name="BaoCao", header=None)
    # Sửa: Lấy danh sách từ hàng thứ 2 (chỉ số 1) trở đi của cột đầu tiên (chỉ số 0)
    project_list = project_df.iloc[1:, 0].tolist()

    wb = load_workbook(excel_path)
    ws = wb["BaoCao"]

    # 1. Đăng nhập
    page.goto("https://qlvh.khaservice.com.vn/login")
    page.locator("input[name='email']").fill("admin@khaservice.com.vn")
    page.locator("input[name='password']").fill("Kha@@123")
    page.locator("button[type='submit']").click()
    page.wait_for_timeout(2000)

    page.goto(f"{base_url}/posts/notification")
    page.locator("//*[@id='root']/div[2]/main/div/div/div[3]/div/div[2]/button").click()
    page.locator("//*[@id='menu-apartment-list-style1']/div[3]/ul/li[6]").click()
    page.wait_for_timeout(2000)

    # 2. Vòng lặp cập nhật danh mục
    for idx, project_val in enumerate(project_list, start=2):  # Sửa: Bắt đầu idx từ 2
        print(f"[{idx}] Project={project_val}")
        logging.error(f"[{idx}] - Project={project_val}")

        page.locator("#combo-box-demo").click()
        page.locator("#combo-box-demo").fill(str(project_val))
        page.locator("#combo-box-demo-option-0").click()

        page.wait_for_timeout(1000)
        rows = page.locator('//*[@id="root"]/div[2]/main/div/div/div[2]/table/tbody/tr')
        notification_count = rows.count()

        logging.error(f"[{idx}] - Project:{project_val} - Tin tuc:{notification_count}")

        ws[f"G{idx}"] = notification_count
    wb.save(excel_path)
    print("Đã ghi xong dữ liệu vào file Excel.")
    page.close()

def test_lay_thong_tin_bai_viet_ngay_cuoi(page: Page):
    excel_path = os.path.join(BASE_DIR, "data.xlsx")
    base_url = "https://qlvh.khaservice.com.vn"
    # Sửa: Đọc file Excel, bỏ qua header để lấy danh sách project từ hàng 2
    project_df = pd.read_excel(excel_path, sheet_name="BaoCao", header=None)
    # Sửa: Lấy danh sách từ hàng thứ 2 (chỉ số 1) trở đi của cột đầu tiên (chỉ số 0)
    project_list = project_df.iloc[1:, 0].tolist()

    wb = load_workbook(excel_path)
    ws = wb["BaoCao"]

    # 1. Đăng nhập
    page.goto("https://qlvh.khaservice.com.vn/login")
    page.locator("input[name='email']").fill("admin@khaservice.com.vn")
    page.locator("input[name='password']").fill("Kha@@123")
    page.locator("button[type='submit']").click()
    page.wait_for_timeout(2000)
    page.goto(f"{base_url}/posts/notification")
    page.wait_for_timeout(2000)

    # # 2. Vòng lặp cập nhật danh mục
    # Vòng lặp
    for idx, project_val in enumerate(project_list, start=2):
        print(f"[{idx}] Project={project_val}")
        logging.error(f"[{idx}] - Project={project_val}")

        page.locator("#combo-box-demo").click()
        page.locator("#combo-box-demo").fill(str(project_val))
        page.locator("#combo-box-demo-option-0").click()
        page.wait_for_timeout(1000)  # Chờ 1 giây để trang cập nhật dữ liệu

        # Khởi tạo giá trị ban đầu là None
        ngay_trang1 = None
        ngay_trang2 = None

        # Lấy giá trị ngày giờ trên trang thông báo
        try:
            page.goto(f"{base_url}/posts/notification")
            locator_thong_bao = page.locator(
                '//*[@id="root"]/div[2]/main/div/div/div[2]/table/tbody/tr[1]/td[8]/div')
            locator_thong_bao.wait_for(timeout=2000)
            ngay_trang1_str = locator_thong_bao.inner_text()
            ngay_trang1 = datetime.strptime(ngay_trang1_str.strip(), '%d/%m/%Y %H:%M')
            logging.error(f"[{idx}] - Ngày trang thông báo: {ngay_trang1_str}")
        except Exception:
            logging.error(f"[{idx}] - Không tìm thấy ngày trên trang thông báo. Bỏ qua.")

        # Lấy giá trị ngày giờ trên trang tin tức
        try:
            page.goto(f"{base_url}/posts/news")
            locator_tin_tuc = page.locator(
                '//*[@id="root"]/div[2]/main/div/div/div[2]/table/tbody/tr[1]/td[8]/div'
            )
            locator_tin_tuc.wait_for(timeout=2000)
            ngay_trang2_str = locator_tin_tuc.inner_text().strip()

            # --- Chỉ lấy phần ngày tháng năm ---
            # cách 1: tách chuỗi
            ngay_trang2_date_str = ngay_trang2_str.split()[0]  # ví dụ '16/09/2025'

            # parse thành datetime để dễ xử lý
            ngay_trang2 = datetime.strptime(ngay_trang2_date_str, '%d/%m/%Y')

            logging.error(f"[{idx}] - Ngày trang tin tức: {ngay_trang2.strftime('%d/%m/%Y')}")
        except Exception:
            logging.error(f"[{idx}] - Không tìm thấy ngày trên trang tin tức. Bỏ qua.")

        # So sánh và ghi vào Excel
        if ngay_trang1 and ngay_trang2:
            # so sánh theo date thôi
            ngay_moi_nhat = max(ngay_trang1, ngay_trang2)
            ws[f"H{idx}"] = ngay_moi_nhat.strftime('%d/%m/%Y')
            logging.error(f"[{idx}] - Ngày mới nhất: {ngay_moi_nhat.strftime('%d/%m/%Y')}")
        elif ngay_trang1:
            ws[f"H{idx}"] = ngay_trang1.strftime('%d/%m/%Y')
            logging.error(f"[{idx}] - Chỉ có ngày trên trang thông báo: {ngay_trang1.strftime('%d/%m/%Y')}")
        elif ngay_trang2:
            ws[f"H{idx}"] = ngay_trang2.strftime('%d/%m/%Y')
            logging.error(f"[{idx}] - Chỉ có ngày trên trang tin tức: {ngay_trang2.strftime('%d/%m/%Y')}")
        else:
            ws[f"H{idx}"] = "Không có dữ liệu"
            logging.error(f"[{idx}] - Không có dữ liệu ngày nào được tìm thấy.")

        wb.save(excel_path)
    print("Đã ghi xong dữ liệu vào file Excel.")
    page.close()

def test_lay_thong_tin_bao_phi_moi_nhat(page: Page):
    excel_path = os.path.join(BASE_DIR, "data.xlsx")
    base_url = "https://qlvh.khaservice.com.vn"
    # Sửa: Đọc file Excel, bỏ qua header để lấy danh sách project từ hàng 2
    project_df = pd.read_excel(excel_path, sheet_name="BaoCao", header=None)
    # Sửa: Lấy danh sách từ hàng thứ 2 (chỉ số 1) trở đi của cột đầu tiên (chỉ số 0)
    project_list = project_df.iloc[1:, 0].tolist()

    wb = load_workbook(excel_path)
    ws = wb["BaoCao"]

    # 1. Đăng nhập
    page.goto("https://qlvh.khaservice.com.vn/login")
    page.locator("input[name='email']").fill("admin@khaservice.com.vn")
    page.locator("input[name='password']").fill("Kha@@123")
    page.locator("button[type='submit']").click()
    page.wait_for_timeout(2000)
    page.goto(f"{base_url}/fee-reports")
    page.wait_for_timeout(2000)

    # # 2. Vòng lặp cập nhật danh mục
    # Vòng lặp
    for idx, project_val in enumerate(project_list, start=2):
        print(f"[{idx}] Project={project_val}")
        logging.error(f"[{idx}] - Project={project_val}")

        page.locator("#combo-box-demo").click()
        page.locator("#combo-box-demo").fill(str(project_val))
        page.locator("#combo-box-demo-option-0").click()
        page.wait_for_timeout(2000)  # Chờ 1 giây để trang cập nhật dữ liệu

        # Lấy giá trị ngày giờ trên trang thông báo
        from datetime import datetime
        # ... (các import khác)

        # Đặt thangmoinhat_text = "" trước try/except để tránh lỗi khi dùng trong except nếu cần
        thangmoinhat_text = ""

        try:
            thangmoinhat_locator = page.locator('//*[@id="root"]/div[2]/main/div/div/div[2]/table/tbody/tr[1]/td[5]/div')  # Đã sửa td[5] thành td[4]
            thangmoinhat_text = thangmoinhat_locator.text_content().strip()
            logging.error(f"[{idx}] - Báo phí mới nhất: {thangmoinhat_text}")
            date_object = datetime.strptime(f"01/{thangmoinhat_text}", '%d/%m/%Y')
            ws[f"I{idx}"] = date_object.strftime('%d/%m/%Y')
            wb.save(excel_path)
        except Exception as e:
            logging.error(f"[{idx}] - Lỗi xảy ra khi xử lý/lưu phí: {e}. Bỏ qua.")
            continue
    print("Đã ghi xong dữ liệu vào file Excel.")
    page.close()


from playwright.sync_api import Page
import pandas as pd
import os
from openpyxl import load_workbook
from datetime import datetime
import logging

# --- VÍ DỤ CẤU HÌNH VÀ HÀM HỖ TRỢ LÙI THÁNG ---
# Giả sử BASE_DIR đã được định nghĩa
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
def get_previous_month(month_str):
    """Chuyển đổi chuỗi MM/YYYY thành đối tượng datetime và lùi lại 1 tháng."""
    try:
        # Giả định tháng hiện tại là 02/2025
        date_obj = datetime.strptime(f"01/{month_str}", '%d/%m/%Y')
        new_month = date_obj.month - 1
        new_year = date_obj.year
        if new_month == 0:
            new_month = 12
            new_year -= 1
        return f"{new_month:02d}/{new_year}"
    except ValueError:
        return None
# --- HÀM CHÍNH TỰ ĐỘNG HÓA ---
def test_xoa_du_lieu_bao_phi_da_thanh_toan(page: Page):
    excel_path = os.path.join(BASE_DIR, "data.xlsx")

    if not os.path.exists(excel_path):
        logging.error(f"Không tìm thấy file Excel tại đường dẫn: {excel_path}")
        return

    # Load dữ liệu
    project_df = pd.read_excel(excel_path, sheet_name="BaoCao1", header=None)
    project_list = project_df.iloc[1:, 0].tolist()
    wb = load_workbook(excel_path)
    ws = wb["BaoCao1"]

    # 🌟 LẤY THÁNG HIỆN TẠI ĐỂ BẮT ĐẦU VÒNG LẶP
    # Định dạng MM/YYYY
    # start_month_str = datetime.now().strftime("%m/%Y")
    start_month_str = datetime.now().strftime("10/2025")
    logging.error(f"Tháng bắt đầu vòng lặp: {start_month_str}")

    # 1. ĐĂNG NHẬP
    page.goto("https://qlvh.khaservice.com.vn/login")
    page.locator("input[name='email']").fill("admin@khaservice.com.vn")
    page.locator("input[name='password']").fill("Kha@@123")
    page.locator("button[type='submit']").click()
    page.wait_for_timeout(2000)

    for idx, project_val in enumerate(project_list, start=2):
        print(f"\n[{idx}] Project={project_val}")
        logging.error(f"[{idx}] - Project={project_val}")

        # 2. CHỌN DỰ ÁN
        try:
            page.locator("#combo-box-demo").click()
            page.locator("#combo-box-demo").fill(str(project_val))
            page.locator("#combo-box-demo-option-0").click()
        except Exception:
            logging.error(f"[{idx}] - Lỗi khi chọn dự án {project_val}. Bỏ qua.")
            continue

        # 3. CHUYỂN ĐẾN TRANG BÁO PHÍ VÀ LẤY THÁNG CŨ NHẤT
        page.locator("//a[@href='/fee-reports']").click()
        page.wait_for_load_state("networkidle")

        # Click để chuyển sang trang cuối (tháng cũ nhất)
        page.locator("xpath=//*[@id='root']/div[2]/main/div/div/div[4]/div/div[1]/nav/ul/li[8]/button").click()
        page.wait_for_timeout(1000)

        try:
            # Lấy tháng cũ nhất từ cột Tháng của hàng đầu tiên (Giả sử td[5])
            thangcunhat_locator = page.locator(
                'xpath=//*[@id="root"]/div[2]/main/div/div/div[2]/table/tbody/tr[1]/td[5]/div')
            thangcunhat = thangcunhat_locator.text_content().strip()
            logging.error(f"[{idx}] - Tháng cũ nhất được tìm thấy: {thangcunhat}")
        except Exception:
            thangcunhat = "01/2000"  # Giá trị mặc định an toàn
            logging.error(f"[{idx}] - Lỗi khi tìm tháng cũ nhất. Đặt mặc định: {thangcunhat}")

        # Quay lại trang đầu
        page.locator("xpath=//*[@id='root']/div[2]/main/div/div/div[4]/div/div[1]/nav/ul/li[2]/button").click()
        page.wait_for_timeout(1000)

        # Click để mở rộng danh sách hiển thị
        page.locator("xpath=//*[@id='root']/div[2]/main/div/div/div[4]/div/div[2]/button").click()
        page.locator("xpath=//*[@id='menu-apartment-list-style1']/div[3]/ul/li[8]").click()
        page.wait_for_timeout(2000)

        current_month_str = start_month_str  # BẮT ĐẦU TỪ THÁNG HIỆN TẠI

        # 4. VÒNG LẶP XÓA NGƯỢC THÁNG
        while True:
            # 🌟 ĐIỀU KIỆN DỪNG VÒNG LẶP (Kiểm tra xem đã lùi quá tháng cũ nhất chưa)
            try:
                date_current = datetime.strptime(f"01/{current_month_str}", '%d/%m/%Y')
                date_oldest = datetime.strptime(f"01/{thangcunhat}", '%d/%m/%Y')

                # Dừng nếu tháng hiện tại nhỏ hơn tháng cũ nhất
                if date_current < date_oldest:
                    logging.error(f"[{idx}] - Đã lùi quá tháng cũ nhất ({thangcunhat}). THOÁT VÒNG LẶP.")
                    # Click để thu nhỏ danh sách hiển thị
                    page.locator("xpath=//*[@id='root']/div[2]/main/div/div/div[4]/div/div[2]/button").click()
                    page.locator("xpath=//*[@id='menu-apartment-list-style1']/div[3]/ul/li[1]").click()
                    page.wait_for_timeout(2000)
                    break
            except ValueError:
                logging.error(f"[{idx}] - Lỗi định dạng tháng trong quá trình so sánh. THOÁT VÒNG LẶP.")
                break

            print(f"[{idx}] Đang xử lý tháng: {current_month_str}")
            logging.error(f"[{idx}] - Đang xử lý tháng: {current_month_str}")

            try:
                # LOCATORs CHUNG
                filter_button = page.locator(
                    "xpath=//*[@id='root']/div[2]/main/div/div/div[1]/div/span/div/div[2]/div/button[2]")
                checkbox_all_locator = page.locator(
                    "xpath=//*[@id='root']/div[2]/main/div/div/div[2]/table/thead/tr/th[1]/span/input")
                delete_button_locator = page.locator(
                    'xpath=//*[@id="root"]/div[2]/main/div/div/div[2]/div[2]/div/div[2]/button')

                # 4.1. MỞ FILTER và ÁP DỤNG LỌC
                filter_button.click()
                page.wait_for_timeout(500)

                # ÁP DỤNG THÁNG MỚI VÀ TRẠNG THÁI 'Đã thanh toán'
                page.locator("xpath=//*[@id='demo-simple-select-helper']").click()
                page.locator("xpath=//*[@data-value='1']").click()
                page.locator("xpath=//*[@placeholder='MM/YYYY']").fill(current_month_str)
                page.keyboard.press("Escape")

                page.wait_for_timeout(3000)  # Đợi dữ liệu load sau khi filter

                # 4.2. KIỂM TRA DỮ LIỆU VÀ XÓA

                if checkbox_all_locator.is_visible():
                    logging.error(
                        f"[{idx}] - TÌM THẤY dữ liệu Đã Thanh Toán cho tháng {current_month_str}. Bắt đầu xóa.")

                    # A. Click chọn tất cả
                    checkbox_all_locator.click()
                    page.wait_for_timeout(500)

                    # B. KIỂM TRA NÚT XÓA VÀ THỰC HIỆN XÓA
                    if delete_button_locator.is_visible():
                        delete_button_locator.click()
                        page.wait_for_timeout(1000)

                        # C. CLICK NÚT XÁC NHẬN TRONG HỘP THOẠI
                        confirm_delete_button = page.locator("xpath=//button[@type='submit']")

                        if confirm_delete_button.is_visible():
                            confirm_delete_button.click()
                            page.wait_for_timeout(3000)
                            logging.error(f"[{idx}] - Đã XÓA thành công dữ liệu tháng {current_month_str}")
                        else:
                            logging.error(f"[{idx}] - LỖI: Không tìm thấy nút XÁC NHẬN XÓA.")
                    else:
                        logging.error(
                            f"[{idx}] - CẢNH BÁO: Đã chọn nhưng nút XÓA không hiển thị. Bỏ qua tháng {current_month_str}.")


                else:
                    logging.error(f"[{idx}] - KHÔNG TÌM THẤY dữ liệu Đã Thanh Toán cho tháng {current_month_str}.")

            except Exception as e:
                # Bắt lỗi chung trong quá trình thao tác hoặc xóa
                logging.error(
                    f"[{idx}] - Lỗi bất ngờ trong vòng lặp tháng {current_month_str}: {e}. Chuyển sang tháng trước.")

            # 5. CHUYỂN SANG THÁNG TRƯỚC
            current_month_str = get_previous_month(current_month_str)
            if current_month_str is None: break

            page.wait_for_timeout(1000)

    page.close()
# --- Toàn bộ Code cập nhật báo cáo ---