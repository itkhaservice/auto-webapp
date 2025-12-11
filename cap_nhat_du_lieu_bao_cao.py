import os
import sys
import subprocess
import logging
import pandas as pd
from playwright.sync_api import sync_playwright, Page
import pytest
from openpyxl import load_workbook
from datetime import datetime
import traceback

# Cấu hình cơ bản
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(BASE_DIR, "data.xlsx")
BASE_URL = "https://qlvh.khaservice.com.vn"


# --- Khởi tạo và Fixtures ---

# Hàm đảm bảo trình duyệt Chromium đã được cài đặt
def ensure_playwright_browsers():
    try:
        from playwright._impl._installer import install
        install("chromium")
    except Exception:
        try:
            subprocess.run(
                [sys.executable, "-m", "playwright", "install", "chromium"],
                check=True, capture_output=True
            )
        except Exception as e:
            print("Không thể tải Chromium:", e)
            sys.exit(1)


@pytest.fixture(scope="session")
def browser_context(show_browser):
    """Fixture khởi tạo trình duyệt và context (scope session)"""
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=not show_browser, args=["--start-maximized"])
        context = browser.new_context(no_viewport=True)
        yield context
        browser.close()


@pytest.fixture(scope="session")
def project_list():
    """Đọc danh sách dự án từ Excel (chỉ đọc một lần)"""
    try:
        project_df = pd.read_excel(EXCEL_PATH, sheet_name="BaoCao", header=None)
        # Lấy danh sách từ hàng thứ 2 (chỉ số 1) trở đi của cột đầu tiên (chỉ số 0)
        return project_df.iloc[1:, 0].tolist()
    except FileNotFoundError:
        logging.error(f"Lỗi: Không tìm thấy file Excel tại đường dẫn: {EXCEL_PATH}")
        sys.exit(1)


@pytest.fixture(scope="session")
def page_login(browser_context: sync_playwright):
    """Fixture đăng nhập và trả về đối tượng page đã đăng nhập (scope session)"""
    page = browser_context.new_page()
    try:
        print("Đang thực hiện đăng nhập...")
        page.goto(f"{BASE_URL}/login")
        # Sử dụng Playwright wait_for_selector thay vì timeout tĩnh
        page.wait_for_selector("input[name='email']")
        page.locator("input[name='email']").fill("admin@khaservice.com.vn")
        page.locator("input[name='password']").fill("Admin@123456")
        page.locator("button[type='submit']").click()
        # Chờ điều hướng đến trang chủ
        page.wait_for_url(f"{BASE_URL}/statistics/overview")
        print("Đăng nhập thành công!")
        return page
    except Exception as e:
        logging.error(f"Lỗi trong quá trình đăng nhập: {e}")
        page.close()
        raise

# --- CÁC HÀM THỰC THI (Execution Functions) ---

def lay_thong_tin_chung(page: Page, project_list: list, ws):
    """Thực thi tác vụ 1: Lấy thông tin chung"""
    print(">>> Bắt đầu Tác vụ 1: Lấy thông tin chung")
    page.goto(f"{BASE_URL}/statistics/overview")
    page.wait_for_selector("#combo-box-demo")

    for idx, project_val in enumerate(project_list, start=2):
        print(f"[{idx}] Tác vụ 1 - Project={project_val}")
        logging.error(f"[{idx}] - Tác vụ 1 - Project={project_val}")

        page.locator("#combo-box-demo").click()
        page.locator("#combo-box-demo").fill(str(project_val))
        page.locator("#combo-box-demo-option-0").click()

        # Chờ dữ liệu tải lên
        page.wait_for_selector('//*[@id="root"]/div[2]/main/div/div/div/div[2]/div/div[1]/p[1]', timeout=10000)

        try:
            tong_can_ho = page.locator('//*[@id="root"]/div[2]/main/div/div/div/div[2]/div/div[1]/p[1]').inner_text()
            tong_cu_dan = page.locator('//*[@id="root"]/div[2]/main/div/div/div/div[3]/div/div[1]/p[1]').inner_text()
            tong_cu_dan_su_dung_app = page.locator(
                '//*[@id="root"]/div[2]/main/div/div/div/div[5]/div/div[1]/p[1]').inner_text()
            tong_can_ho_su_dung_app = page.locator(
                '//*[@id="root"]/div[2]/main/div/div/div/div[6]/div/div[1]/p[1]').inner_text()

            # Ghi vào Excel
            ws[f"B{idx}"] = tong_can_ho
            ws[f"C{idx}"] = tong_cu_dan
            ws[f"D{idx}"] = tong_cu_dan_su_dung_app
            ws[f"E{idx}"] = tong_can_ho_su_dung_app
        except Exception as e:
            logging.error(f"[{idx}] Lỗi khi lấy thông tin chung: {e}")
            ws[f"B{idx}"] = "Lỗi"
    print("<<< Hoàn thành Tác vụ 1.")


def lay_so_luong_tin_tuc(page: Page, project_list: list, ws):
    """Thực thi tác vụ 2: Lấy số lượng tin tức"""
    print(">>> Bắt đầu Tác vụ 2: Lấy số lượng tin tức")
    page.goto(f"{BASE_URL}/posts/news")
    # Đặt lại bộ lọc để đảm bảo hiển thị hết dữ liệu (ví dụ: hiển thị 100 dòng)
    page.locator("button:has-text('10')").click()
    page.locator("li:has-text('100')").click()
    page.wait_for_timeout(1000)  # Chờ cập nhật danh sách

    for idx, project_val in enumerate(project_list, start=2):
        print(f"[{idx}] Tác vụ 2 - Project={project_val}")

        page.locator("#combo-box-demo").click()
        page.locator("#combo-box-demo").fill(str(project_val))
        page.locator("#combo-box-demo-option-0").click()
        page.wait_for_timeout(1000)  # Chờ thay đổi dự án

        try:
            rows = page.locator('//*[@id="root"]/div[2]/main/div/div/div[2]/table/tbody/tr')
            tin_tuc_count = rows.count()
            ws[f"F{idx}"] = tin_tuc_count
            logging.error(f"[{idx}] - Project:{project_val} - Tin tuc:{tin_tuc_count}")
        except Exception as e:
            logging.error(f"[{idx}] Lỗi khi lấy số lượng tin tức: {e}")
            ws[f"F{idx}"] = "Lỗi"
    print("<<< Hoàn thành Tác vụ 2.")


def lay_so_luong_thong_bao(page: Page, project_list: list, ws):
    """Thực thi tác vụ 3: Lấy số lượng thông báo"""
    print(">>> Bắt đầu Tác vụ 3: Lấy số lượng thông báo")
    page.goto(f"{BASE_URL}/posts/notification")
    # Đặt lại bộ lọc để đảm bảo hiển thị hết dữ liệu (ví dụ: hiển thị 100 dòng)
    page.locator("button:has-text('10')").click()
    page.locator("li:has-text('100')").click()
    page.wait_for_timeout(1000)  # Chờ cập nhật danh sách

    for idx, project_val in enumerate(project_list, start=2):
        print(f"[{idx}] Tác vụ 3 - Project={project_val}")

        page.locator("#combo-box-demo").click()
        page.locator("#combo-box-demo").fill(str(project_val))
        page.locator("#combo-box-demo-option-0").click()
        page.wait_for_timeout(1000)  # Chờ thay đổi dự án

        try:
            rows = page.locator('//*[@id="root"]/div[2]/main/div/div/div[2]/table/tbody/tr')
            notification_count = rows.count()
            ws[f"G{idx}"] = notification_count
            logging.error(f"[{idx}] - Project:{project_val} - Thong bao:{notification_count}")
        except Exception as e:
            logging.error(f"[{idx}] Lỗi khi lấy số lượng thông báo: {e}")
            ws[f"G{idx}"] = "Lỗi"
    print("<<< Hoàn thành Tác vụ 3.")


def lay_thong_tin_bai_viet_ngay_cuoi(page: Page, project_list: list, ws):
    """Thực thi tác vụ 4: Lấy ngày cuối cùng đăng bài"""
    print(">>> Bắt đầu Tác vụ 4: Lấy ngày cuối cùng đăng bài")

    for idx, project_val in enumerate(project_list, start=2):
        print(f"[{idx}] Tác vụ 4 - Project={project_val}")

        # Chọn dự án (chọn một lần và giữ trạng thái cho các lần goto sau)
        page.locator("#combo-box-demo").click()
        page.locator("#combo-box-demo").fill(str(project_val))
        page.locator("#combo-box-demo-option-0").click()
        page.wait_for_timeout(500)

        ngay_trang1 = None  # Thông báo
        ngay_trang2 = None  # Tin tức

        # 1. Lấy ngày trên trang Thông báo
        try:
            page.goto(f"{BASE_URL}/posts/notification")
            locator_thong_bao = page.locator(
                '//*[@id="root"]/div[2]/main/div/div/div[2]/table/tbody/tr[1]/td[8]/div')
            locator_thong_bao.wait_for(timeout=3000)
            ngay_trang1_str = locator_thong_bao.inner_text().strip()
            # Giả định: 'dd/mm/yyyy hh:mm'
            ngay_trang1 = datetime.strptime(ngay_trang1_str, '%d/%m/%Y %H:%M').date()
            logging.error(f"[{idx}] - Ngày thông báo: {ngay_trang1}")
        except Exception:
            logging.error(f"[{idx}] - Không tìm thấy ngày trên trang thông báo. Bỏ qua.")

        # 2. Lấy ngày trên trang Tin tức
        try:
            page.goto(f"{BASE_URL}/posts/news")
            locator_tin_tuc = page.locator(
                '//*[@id="root"]/div[2]/main/div/div/div[2]/table/tbody/tr[1]/td[8]/div'
            )
            locator_tin_tuc.wait_for(timeout=3000)
            ngay_trang2_str = locator_tin_tuc.inner_text().strip()
            # Giả định: 'dd/mm/yyyy hh:mm' -> chỉ lấy ngày
            ngay_trang2_date_str = ngay_trang2_str.split()[0]
            ngay_trang2 = datetime.strptime(ngay_trang2_date_str, '%d/%m/%Y').date()
            logging.error(f"[{idx}] - Ngày tin tức: {ngay_trang2}")
        except Exception:
            logging.error(f"[{idx}] - Không tìm thấy ngày trên trang tin tức. Bỏ qua.")

        # 3. So sánh và ghi vào Excel
        if ngay_trang1 and ngay_trang2:
            ngay_moi_nhat = max(ngay_trang1, ngay_trang2)
            ws[f"H{idx}"] = ngay_moi_nhat.strftime('%d/%m/%Y')
        elif ngay_trang1:
            ws[f"H{idx}"] = ngay_trang1.strftime('%d/%m/%Y')
        elif ngay_trang2:
            ws[f"H{idx}"] = ngay_trang2.strftime('%d/%m/%Y')
        else:
            ws[f"H{idx}"] = "Không có dữ liệu"

    print("<<< Hoàn thành Tác vụ 4.")


# --- HÀM CHẠY CHƯƠNG TRÌNH (Main Execution) ---

def run_test(show_browser, selected_tasks):
    """Thực thi các tác vụ đã chọn bằng cách gọi pytest"""

    # Gán các hàm thực thi vào dictionary
    task_map = {
        1: lay_thong_tin_chung,
        2: lay_so_luong_tin_tuc,
        3: lay_so_luong_thong_bao,
        4: lay_thong_tin_bai_viet_ngay_cuoi,
    }

    print("\n--- BẮT ĐẦU THU THẬP DỮ LIỆU ---")

    try:
        # Tải workbook và danh sách project một lần
        wb = load_workbook(EXCEL_PATH)
        ws = wb["BaoCao"]

        # Chạy Pytest để tận dụng các fixtures scope="session" (đăng nhập 1 lần)
        # Chúng ta sẽ dùng một hàm wrapper để gọi pytest và nhận các fixtures
        class RunWrapper:
            """Một lớp dummy để Pytest có thể inject fixtures"""

            def __init__(self, show_browser):
                self.show_browser = show_browser

            @pytest.mark.parametrize('show_browser', [show_browser])
            def test_wrapper(self, page_login: Page, project_list: list):
                """Hàm wrapper Pytest thực sự gọi các execution functions"""
                # Bắt đầu vòng lặp thực thi các tác vụ đã chọn
                for task_id in selected_tasks:
                    if task_id in task_map:
                        task_map[task_id](page_login, project_list, ws)

        # Chạy Pytest
        pytest.main([
            __file__,
            '--verbose',
            '--capture=no',
            '--tb=no',
            f'--browser={show_browser}'  # Tận dụng argument để truyền vào fixture
        ], plugins=[RunWrapper(show_browser)])

        # Lưu file sau khi tất cả các tác vụ đã hoàn tất
        wb.save(EXCEL_PATH)
        print("--- Đã ghi xong dữ liệu vào file Excel. ---")

    except Exception as e:
        logging.error(f"Lỗi chính trong quá trình thực thi: {e}")
        traceback.print_exc()


# --- HÀM MAIN CỦA CHƯƠNG TRÌNH ---

def main():
    # ... (giữ nguyên hàm main để xử lý menu người dùng) ...
    try:
        ensure_playwright_browsers()

        print("\n--- CHƯƠNG TRÌNH LẤY DỮ LIỆU BÁO CÁO TỰ ĐỘNG ---")
        print("▶ Bạn có muốn hiển thị trình duyệt không?")
        print("▶ Lưu ý: Chọn 'Y' sẽ giúp bạn theo dõi quá trình, nhưng có thể sẽ chậm hơn.")
        choice_browser = input("Nhập (Y/N) để chọn chế độ hiển thị: ").strip().lower()
        show_browser = True if choice_browser == 'y' else False

        selected_tasks = []
        while True:
            print("\n======================")
            print("\n--- MENU CHỨC NĂNG ---")
            print("1. Lấy thông tin chung (Căn hộ, Cư dân, App)")
            print("2. Lấy số lượng tin tức")
            print("3. Lấy số lượng thông báo")
            print("4. Lấy ngày cuối cùng đăng bài")
            print("5. Chạy tất cả các tác vụ (1, 2, 3, 4)")
            print("0. Thoát chương trình")
            print("\n======================")
            task_choice = input("▶ Vui lòng chọn chức năng (1-5, hoặc 0 để thoát): ").strip()
            print("\n======================")
            if task_choice == '0':
                print("▶ Đang thoát chương trình... Tạm biệt!")
                break
            elif task_choice == '5':
                selected_tasks = [1, 2, 3, 4]
                run_test(show_browser, selected_tasks)
                break
            elif task_choice.isdigit() and int(task_choice) in range(1, 5):
                selected_tasks = [int(task_choice)]
                run_test(show_browser, selected_tasks)
            else:
                print("▶ Lựa chọn không hợp lệ. Vui lòng nhập lại...")

    except Exception as e:
        logging.error("▶ Chương trình bị lỗi ở mục chính...")
        traceback.print_exc()
    finally:
        input("\nNhấn Enter để thoát chương trình...")


if __name__ == "__main__":
    main()